import os
import gitlab
import logging
import shutil
import pathlib
import time
import yaml
from openpyxl import Workbook
from openpyxl import load_workbook
from git import Repo

start_time = time.time()

# 读取配置文件
with open('config.yml') as file:
    config = yaml.safe_load(file)

# 1. 设置 GitLab API 访问令牌
GITLAB_API_ACCESS_TOKEN = config['gitlab_api_access_token']

# 2. 设置 GitLab 群组 ID
GROUP_ID = config['group_id']

# 3. 设置查找的接口名称
API_TO_SEARCH = config['api_to_search']

# 4. 设置 GitLab API 客户端
gl = gitlab.Gitlab(config['gitlab_api_url'], private_token=GITLAB_API_ACCESS_TOKEN)

# 5. 设置遍历的分支数
BRANCH_LIMIT = config['branch_limit']

# 6. 设置是否为测试模式
TEST_MODE = config['test_mode']

# 开始前先删除文件
files_to_delete = ['./match.log', './branch.log', './err.log', './match.xlsx']

for file in files_to_delete:
    if os.path.exists(file):
        os.remove(file)
        print(f"文件 {file} 已删除")
    else:
        print(f"文件 {file} 不存在")

# 尝试打开 Excel 文件以检查是否可以读写
try:
    workbook = load_workbook(filename='match.xlsx')
    worksheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "匹配结果"
    worksheet.append(["仓库名", "分支名", "文件路径", "文件行数"])
    worksheet = workbook.active
except IOError as e:
    logging.error(f"无法打开 Excel 文件：{e}")
    end_time = time.time()
    elapsed_time = end_time - start_time

    print(f"程序总共耗时 {elapsed_time:.2f} 秒")
    exit()

# 创建一个新的 Excel 工作簿
if not worksheet.cell(row=1, column=1).value:
    worksheet.title = "匹配结果"
    worksheet.append(["仓库名", "分支名", "文件路径", "文件行数"])

# 配置 output.log 记录器
logging.basicConfig(filename='output.log', level=logging.INFO, format='%(message)s')

# 创建一个新的日志记录器并配置 match.log
match_logger = logging.getLogger('match')
match_handler = logging.FileHandler('match.log')
match_handler.setFormatter(logging.Formatter('%(message)s'))
match_logger.addHandler(match_handler)
match_logger.setLevel(logging.INFO)

# 创建一个新的日志记录器并配置 branch.log
branch_logger = logging.getLogger('branch')
branch_handler = logging.FileHandler('branch.log')
branch_handler.setFormatter(logging.Formatter('%(message)s'))
branch_logger.addHandler(branch_handler)
branch_logger.setLevel(logging.INFO)

err_logger = logging.getLogger('err')
err_handler = logging.FileHandler('err.log')
err_handler.setFormatter(logging.Formatter('%(message)s'))
err_logger.addHandler(err_handler)
err_logger.setLevel(logging.INFO)


# 定义一个函数，用于从给定的仓库和分支中获取所有的 JS 文件
def get_js_files_from_repo(local_repo_path, path=""):
    js_files = []

    for root, dirs, files in os.walk(os.path.join(local_repo_path, path)):
        # 排除 .git 目录和其他隐藏文件等
        dirs[:] = [d for d in dirs if not d.startswith('.')]
        files[:] = [f for f in files if not f.startswith('.')]

        for file in files:
            if file.endswith('.js'):
                js_file_path = os.path.relpath(os.path.join(root, file), local_repo_path)
                js_files.append(js_file_path)

    return js_files


# 定义一个函数，用于在文件内容中搜索给定的 API 名称
def search_api_in_file(file_content, api_name):
    matched_lines = []

    for line_number, line in enumerate(file_content.split('\n')):
        if api_name in line:
            matched_lines.append((line_number + 1, line))

    return matched_lines


def main():
    group = gl.groups.get(GROUP_ID)
    projects = group.projects.list(all=True)

    # 添加本次任务执行时间
    logging.info(f"本次任务执行时间: {time.strftime('%Y-%m-%d %H:%M:%S')}")

    for project in projects:
        print(project.name)

    # 测试模式下只遍历第一个仓库
    if TEST_MODE:
        projects = projects[:1]

    for project in projects:
        logging.info("==== 大分割线 ====")
        logging.info(f"正在处理项目：{project.name}")
        repo = gl.projects.get(project.id)
        branches = repo.branches.list(get_all=True)
        all_branch_names = [branch.name for branch in branches]

        # 只遍历指定的分支数
        selected_branches = all_branch_names[:BRANCH_LIMIT]

        for branch in selected_branches:
            logging.info(f"----- 小分割线 -----")
            logging.info(f"正在处理分支：{branch}")

            # 检查目录是否存在，如果存在则删除
            local_repo_path = f"{project.name}-{branch}"
            local_repo_pathlib = pathlib.Path(local_repo_path)
            if local_repo_pathlib.exists():
                shutil.rmtree(local_repo_path, ignore_errors=True)

            # 克隆仓库到本地临时目录
            local_repo_path = f"{project.name}-{branch}"
            Repo.clone_from(repo.http_url_to_repo, local_repo_path, branch=branch)

            try:
                # 从本地仓库获取 JS 文件列表
                js_files = get_js_files_from_repo(local_repo_path)

                # 添加分支信息到 branch.log
                local_repo = Repo(local_repo_path)
                branch_commit = local_repo.commit(branch)
                commit_time = time.strftime('%Y-%m-%d %H:%M:%S', time.gmtime(branch_commit.committed_date))
                branch_info = f"{repo.name}-{branch}-{branch_commit.hexsha}-{commit_time}-{branch_commit.author}"
                branch_logger.info(branch_info)

                for js_file_path in js_files:
                    logging.info(f"正在处理 JS 文件：{os.path.abspath(os.path.join(local_repo_path, js_file_path))}")
                    print(f"正在处理 JS 文件：{os.path.abspath(os.path.join(local_repo_path, js_file_path))}")

                    with open(os.path.join(local_repo_path, js_file_path), 'r', encoding='utf-8') as file:
                        # 处理文件内容
                        content = file.read()

                        matched_lines = search_api_in_file(content, API_TO_SEARCH)
                        if matched_lines:
                            for line_number, line in matched_lines:
                                match_info = f"{repo.name}-{branch}-{js_file_path}-{line_number}-{line.strip()}"
                                logging.info(match_info)
                                match_logger.info(match_info)

                                # 将匹配结果添加到 Excel 工作表
                                worksheet.append([repo.name, branch, js_file_path, line_number, line])

            except Exception as e:
                err_logger.error(f"项目：{project.name}，处理分支 {branch} 时出错：{e}")
            finally:
                # 删除本地仓库临时目录
                shutil.rmtree(local_repo_path, ignore_errors=True)

            # 测试模式下只遍历第一个仓库
            if TEST_MODE:
                break

        # 保存 Excel 工作簿到文件
        workbook.save("match.xlsx")

    end_time = time.time()
    elapsed_time = end_time - start_time

    logging.info(f"程序总共耗时 {elapsed_time:.4f} 秒")
    print(f"程序总共耗时 {elapsed_time:.4f} 秒")


if __name__ == '__main__':
    main()
