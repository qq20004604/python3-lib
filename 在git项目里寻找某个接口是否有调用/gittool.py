import os
import gitlab
import logging
import shutil
import pathlib
import time
import yaml
from openpyxl import Workbook
from openpyxl import load_workbook
from git import Repo
import re

import ctypes
import json


class GitTool:
    def __init__(self):
        # 搜索字符串
        self.search_string = None
        # 配置
        self.config = None
        # gitlab api的客户端
        self.gl = None
        # excel
        self.worksheet = None
        # 日志句柄
        self.match_logger = None
        self.branch_logger = None
        self.err_logger = None
        self.search_lib = None

    # 读取配置
    def read_config(self):
        # 读取配置文件
        with open('config.yml') as file:
            config = yaml.safe_load(file)

        self.config = config
        # 3. 设置查找的接口名称
        self.search_string = self.config['api_to_search']
        # 4. 设置 GitLab API 客户端
        self.gl = gitlab.Gitlab(config['gitlab_api_url'], private_token=self.config['gitlab_api_access_token'])

        # 检查配置
        if self.config['model'] not in ['group', 'repository', 'group']:
            print("model 模式错误，只能是 group 群组模式，repository 单项目模式，repositories 多项目模式")
            exit()

        return config

    # 删除历史文件
    def remove_oldfile(self):
        # 开始前先删除文件
        files_to_delete = ['./match.log', './branch.log', './err.log', './match.xlsx']

        for file in files_to_delete:
            if os.path.exists(file):
                os.remove(file)
                print(f"文件 {file} 已删除")
            else:
                print(f"文件 {file} 不存在")

    # 检查excel
    def open_excel(self):
        # 尝试打开 Excel 文件以检查是否可以读写
        try:
            workbook = load_workbook(filename='match.xlsx')
            self.worksheet = workbook.active
            # 清空之前的内容
            for row in self.worksheet.iter_rows(min_row=2, min_col=1):
                for cell in row:
                    cell.value = None
        except FileNotFoundError:
            workbook = Workbook()
            self.worksheet = workbook.active
            self.worksheet.title = "匹配结果"
            self.worksheet.append(["仓库名", "分支名", "文件路径", "文件行数", "该行内容"])
            self.worksheet = workbook.active
        except IOError as e:
            logging.error(f"无法打开 Excel 文件：{e}")
            exit()

        # 创建一个新的 Excel 工作簿
        if not self.worksheet.cell(row=1, column=1).value:
            self.worksheet.title = "匹配结果"
            self.worksheet.append(["仓库名", "分支名", "文件路径", "文件行数"])

    # 日志设置
    def set_log(self):
        # 配置 output.log 记录器
        logging.basicConfig(filename='output.log', level=logging.INFO, format='%(message)s')

        # 创建一个新的日志记录器并配置 match.log
        self.match_logger = logging.getLogger('match')
        match_handler = logging.FileHandler('match.log')
        match_handler.setFormatter(logging.Formatter('%(message)s'))
        self.match_logger.addHandler(match_handler)
        self.match_logger.setLevel(logging.INFO)

        # 创建一个新的日志记录器并配置 branch.log
        self.branch_logger = logging.getLogger('branch')
        branch_handler = logging.FileHandler('branch.log')
        branch_handler.setFormatter(logging.Formatter('%(message)s'))
        self.branch_logger.addHandler(branch_handler)
        self.branch_logger.setLevel(logging.INFO)

        self.err_logger = logging.getLogger('err')
        err_handler = logging.FileHandler('err.log')
        err_handler.setFormatter(logging.Formatter('%(message)s'))
        self.err_logger.addHandler(err_handler)
        self.err_logger.setLevel(logging.INFO)

    # 遍历所有符合的文件，然后返回文件名
    # 入参是本地地址
    def find_files_by_match_name(self, local_repo_path, path=""):
        js_files = []

        for root, dirs, files in os.walk(os.path.join(local_repo_path, path)):
            # 排除 .git 目录和其他隐藏文件等
            dirs[:] = [d for d in dirs if not d.startswith('.')]
            files[:] = [f for f in files if not f.startswith('.')]

            # 全文件匹配
            if self.config['file_match']['type'] == "all":
                for file in files:
                    js_file_path = os.path.relpath(os.path.join(root, file), local_repo_path)
                    js_files.append(js_file_path)
            # 字符串匹配文件名
            elif self.config['file_match']['type'] == "normal":
                for file in files:
                    # 如果这些某个文件以这个为结尾
                    if any(file in type for type in self.config['file_match']['file_type']):
                        js_file_path = os.path.relpath(os.path.join(root, file), local_repo_path)
                        js_files.append(js_file_path)
            # 后缀名匹配
            elif self.config['file_match']['type'] == "ext":
                for file in files:
                    # 如果这些某个文件以这个为结尾
                    if any(file.endswith(ext) for ext in self.config['file_match']['file_type']):
                        js_file_path = os.path.relpath(os.path.join(root, file), local_repo_path)
                        js_files.append(js_file_path)
            # 正则匹配
            elif self.config['file_match']['type'] == "regexp":
                for file in files:
                    regex = re.compile()
                    if len(regex.findall(file)) > 0:
                        js_file_path = os.path.relpath(os.path.join(root, file), local_repo_path)
                        js_files.append(js_file_path)
            else:
                self.err_logger.info("未选择文件匹配模式，程序退出")
                exit()

        return js_files

    # 定义一个函数，用于在文件内容中搜索给定的【字符串】
    def search_string_in_file_by_python(self, file_path, search_string):
        res = {
            "MatchedLines": [],
            "Error": "",
        }

        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                # 处理文件内容
                file_content = file.read()
                for line_number, line in enumerate(file_content.split('\n')):
                    if search_string in line:
                        item = {
                            "LineNumber": line_number + 1,
                            "Line": line
                        }
                        res['MatchedLines'].append(item)
        except FileNotFoundError:
            err = f"文件未找到：{file_path}"
            res['Error'] = err
            print(err)
        except PermissionError:
            err = f"没有权限读取文件：{file_path}"
            res['Error'] = err
            print(err)
        except IOError as e:
            err = f"读取文件时发生错误：{e}"
            res['Error'] = err
            print(err)
        finally:
            return res

    # 入参：文件路径，搜索字符串
    # 出参：一个对象。
    #   result.get('Error') 为 True 表示有报错信息，通过 result['Error'] 获取报错信息
    #   result['MatchedLines'] 为匹配成功的数据，len=0说明没有任何被匹配成功
    #   result['MatchedLines'][0] 示例结果 {'LineNumber': 353, 'Line': '仓库名-分支名-文件相对路径-行数-该行代码内容'}
    def search_string_in_file_by_go(self, file_path, search_string):
        # 如果没有初始化 go 库，则初始化
        if self.search_lib is None:
            self.search_lib = ctypes.CDLL(os.path.abspath("./lib/search.so"))
            self.search_lib.searchStringInFileWrapper.argtypes = [ctypes.c_char_p, ctypes.c_char_p]
            self.search_lib.searchStringInFileWrapper.restype = ctypes.c_char_p
        # 将字符串转义
        c_file_path = ctypes.c_char_p(file_path.encode('utf-8'))
        c_search_string = ctypes.c_char_p(search_string.encode('utf-8'))
        # 调用go库
        c_result = self.search_lib.searchStringInFileWrapper(c_file_path, c_search_string)
        # 拉取返回结果
        result = json.loads(ctypes.string_at(c_result).decode('utf-8'))
        return result


if __name__ == '__main__':
    gt = GitTool()
    # # 设置日志
    # gt.set_log()
    # # 读取配置
    # gt.read_config()
    # # 删除历史文件
    # gt.remove_oldfile()
    # # 启动 excel
    # gt.open_excel()
    # with open('config.yml', 'r') as file:
    #     data = yaml.safe_load(file)
    #     print(data['file_type'])

    start_time = time.time()

    file_path = "./output.log"
    search_string = "approvalComments/"

    # hello_lib = ctypes.CDLL(os.path.abspath("./lib/hello.so"))
    # hello_lib.Hello.argtypes = [ctypes.c_char_p]
    # hello_lib.Hello.restype = None
    # name = "John Doe"
    # hello_lib.Hello(name.encode('utf-8'))
    #
    # search_lib = ctypes.CDLL(os.path.abspath("./lib/search.so"))
    # search_lib.searchStringInFileWrapper.argtypes = [ctypes.c_char_p, ctypes.c_char_p]
    # search_lib.searchStringInFileWrapper.restype = ctypes.c_char_p
    #
    file_path = './output.log'
    search_string = 'approvalComments'
    result = gt.search_string_in_file_by_python(file_path, search_string)

    end_time = time.time()
    elapsed_time = end_time - start_time

    print(f"程序总共耗时 {elapsed_time:.4f} 秒")
    print(result)
    print(gt.search_string_in_file_by_go(file_path, search_string))
    # print(result['Error'])
    # if result.get('Error'):
    #     print(result['Error'])
    # else:
    #     print(result['MatchedLines'])
    # c_file_path = ctypes.c_char_p(file_path.encode('utf-8'))
    # c_search_string = ctypes.c_char_p(search_string.encode('utf-8'))
    #
    # c_result = search_lib.searchStringInFileWrapper(c_file_path, c_search_string)
    # result = json.loads(ctypes.string_at(c_result).decode('utf-8'))
    # print(result)
